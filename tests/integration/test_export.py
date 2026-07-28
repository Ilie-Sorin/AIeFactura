"""Export: structura de foldere ZIP + Excel (cap. 3, 9)."""

import zipfile
from io import BytesIO
from pathlib import Path

import openpyxl

from app.services.export import (
    build_bulk_export_zip,
    build_document_export_zip,
    export_registry_to_excel,
)
from app.services.ingest import IngestFile, finish_batch, ingest_file, start_batch

FIXTURES = Path(__file__).resolve().parent.parent / "fixtures"


def _importa(session, nume: str):
    batch = start_batch(session, tip="scan_local", sursa="test")
    rezultat = ingest_file(session, batch, IngestFile((FIXTURES / nume).read_bytes(), nume))
    finish_batch(session, batch)
    return rezultat


def test_document_export_zip_contains_xml_and_metadata(db_session_commit):
    session = db_session_commit
    r = _importa(session, "factura_normala.xml")
    session.commit()

    from app.models.document import Invoice

    invoice = session.get(Invoice, r.invoice_id)

    continut = build_document_export_zip(session, invoice)
    zf = zipfile.ZipFile(BytesIO(continut))
    nume = zf.namelist()

    prefix = f"143981007/2026/Primite/2026-03/factura_{invoice.id}"
    assert f"{prefix}/factura.xml" in nume
    assert f"{prefix}/metadata.json" in nume

    xml_continut = zf.read(f"{prefix}/factura.xml")
    assert b"0001234" in xml_continut

    import json

    metadata = json.loads(zf.read(f"{prefix}/metadata.json"))
    assert metadata["numar_brut"] == "0001234"
    assert metadata["cif_emitent"] == "185472901"
    assert metadata["total_document"] == "18599.70"


def test_bulk_export_zip_contains_all_invoices(db_session_commit):
    session = db_session_commit
    r1 = _importa(session, "factura_normala.xml")
    r2 = _importa(session, "factura_storno.xml")
    session.commit()

    from app.models.document import Invoice

    invoices = [session.get(Invoice, r1.invoice_id), session.get(Invoice, r2.invoice_id)]
    continut = build_bulk_export_zip(session, invoices)
    zf = zipfile.ZipFile(BytesIO(continut))
    nume = zf.namelist()

    assert any(f"factura_{invoices[0].id}/factura.xml" in n for n in nume)
    assert any(f"factura_{invoices[1].id}/factura.xml" in n for n in nume)
    assert all(n.startswith("Export/") for n in nume)


def test_excel_export_has_expected_columns_and_values(db_session_commit):
    session = db_session_commit
    r = _importa(session, "factura_normala.xml")
    session.commit()

    from app.models.document import Invoice

    invoice = session.get(Invoice, r.invoice_id)
    continut = export_registry_to_excel([invoice])

    wb = openpyxl.load_workbook(BytesIO(continut))
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "Număr" in header
    assert "Total document" in header

    rand = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    linie = dict(zip(header, rand))
    assert linie["Număr"] == "0001234"
    assert linie["CIF emitent"] == "185472901"
    assert round(linie["Total document"], 2) == 18599.70
