"""Export (cap. 3, 9): structura de foldere devine FORMAT DE SCHIMB, nu format
de stocare — reconstituită la cerere ca ZIP, plus export Excel al registrului.

```
Export/
└── {CIF propriu}/{an}/{Primite|Emise}/{an-luna}/factura_{id}/
    ├── original.zip        (dacă documentul a sosit într-o arhivă)
    ├── factura.xml
    ├── semnatura.xml       (dacă a fost prezentă)
    └── metadata.json
```
"""

from __future__ import annotations

import json
import zipfile
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.document import Invoice
from app.models.ingestion import InvoiceSourceLink, SourceObject


def _own_cif_for_invoice(invoice: Invoice) -> str:
    if invoice.directie == "iesire":
        return invoice.cif_emitent
    return invoice.cif_beneficiar or invoice.cif_emitent


def export_relative_dir(invoice: Invoice) -> str:
    cif = _own_cif_for_invoice(invoice)
    an = invoice.data_emitere.year
    directie_folder = "Emise" if invoice.directie == "iesire" else "Primite"
    luna = invoice.data_emitere.strftime("%Y-%m")
    return f"{cif}/{an}/{directie_folder}/{luna}/factura_{invoice.id}"


def _metadata_json(invoice: Invoice, xml_source: SourceObject | None) -> str:
    metadata = {
        "id": invoice.id,
        "directie": invoice.directie,
        "cif_emitent": invoice.cif_emitent,
        "cif_beneficiar": invoice.cif_beneficiar,
        "numar_brut": invoice.numar_brut,
        "numar_normalizat": invoice.numar_normalizat,
        "data_emitere": invoice.data_emitere.isoformat(),
        "data_scadenta": invoice.data_scadenta.isoformat() if invoice.data_scadenta else None,
        "moneda": invoice.moneda,
        "total_fara_tva": str(invoice.total_fara_tva),
        "total_tva": str(invoice.total_tva),
        "total_document": str(invoice.total_document),
        "stare": invoice.stare,
        "versiune_cius": invoice.versiune_cius,
        "sha256_xml": xml_source.sha256 if xml_source else None,
    }
    return json.dumps(metadata, ensure_ascii=False, indent=2)


def _write_invoice_bundle(zf: zipfile.ZipFile, session: Session, invoice: Invoice, prefix: str) -> None:
    xml_source = session.get(SourceObject, invoice.source_object_id)
    if xml_source is not None:
        zf.writestr(f"{prefix}/factura.xml", xml_source.continut)

    legaturi = session.scalars(
        select(InvoiceSourceLink).where(InvoiceSourceLink.invoice_id == invoice.id)
    ).all()
    for legatura in legaturi:
        obiect = session.get(SourceObject, legatura.source_object_id)
        if obiect is None:
            continue
        if obiect.tip == "zip":
            zf.writestr(f"{prefix}/original.zip", obiect.continut)
        elif obiect.tip == "xml_semnatura":
            zf.writestr(f"{prefix}/semnatura.xml", obiect.continut)

    zf.writestr(f"{prefix}/metadata.json", _metadata_json(invoice, xml_source))


def build_document_export_zip(session: Session, invoice: Invoice) -> bytes:
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        _write_invoice_bundle(zf, session, invoice, export_relative_dir(invoice))
    return buffer.getvalue()


def build_bulk_export_zip(session: Session, invoices: list[Invoice]) -> bytes:
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for invoice in invoices:
            _write_invoice_bundle(zf, session, invoice, f"Export/{export_relative_dir(invoice)}")
    return buffer.getvalue()


REGISTRY_COLUMNS = (
    ("Direcție", "directie"),
    ("Număr", "numar_brut"),
    ("Număr normalizat", "numar_normalizat"),
    ("CIF emitent", "cif_emitent"),
    ("CIF beneficiar", "cif_beneficiar"),
    ("Dată emitere", "data_emitere"),
    ("Total fără TVA", "total_fara_tva"),
    ("Total TVA", "total_tva"),
    ("Total document", "total_document"),
    ("Monedă", "moneda"),
    ("Stare", "stare"),
    ("Contract", "nr_contract"),
    ("Comandă", "nr_comanda"),
)


def export_registry_to_excel(invoices: list[Invoice]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Registru"
    ws.append([titlu for titlu, _ in REGISTRY_COLUMNS])

    for invoice in invoices:
        rand = []
        for _, camp in REGISTRY_COLUMNS:
            valoare = getattr(invoice, camp)
            if isinstance(valoare, Decimal):
                valoare = float(valoare)
            rand.append(valoare)
        ws.append(rand)

    for idx, (titlu, _) in enumerate(REGISTRY_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(12, len(titlu) + 2)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
